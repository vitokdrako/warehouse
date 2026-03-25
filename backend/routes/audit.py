"""
Audit Cabinet API - Кабінет переобліку
"""
from fastapi import APIRouter, HTTPException, Depends, Body, UploadFile
import fastapi
from typing import List, Optional
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_, desc, text
from datetime import datetime, timedelta
import uuid
import html

from database import get_db as get_oc_db  # OpenCart DB (for fallback)
from database_rentalhub import get_rh_db  # RentalHub DB (primary)
from utils.image_helper import normalize_image_url
from models_sqlalchemy import (
    OpenCartProduct,
    OpenCartProductDescription,
    OpenCartProductToCategory,
    OpenCartCategory,
    OpenCartCategoryDescription,
    DecorProductExtended,
    DecorProductCatalog,
    DecorInventoryItem,
    DecorProductHistory,
    DecorOrder,
    DecorIssueCard,
    DecorDamage,
    DecorDamageItem,
    FinanceTransaction
)

router = APIRouter(prefix="/api/audit", tags=["audit"])


# Helper to calculate days from last audit
def days_from_date(date_obj):
    if not date_obj:
        return 999
    try:
        from datetime import date as date_type
        
        if isinstance(date_obj, str):
            audit_date = datetime.strptime(date_obj, "%Y-%m-%d").date()
        elif isinstance(date_obj, date_type):
            audit_date = date_obj
        elif isinstance(date_obj, datetime):
            audit_date = date_obj.date()
        else:
            return 999
            
        today = datetime.now().date()
        delta = (today - audit_date).days
        return delta
    except Exception as e:
        print(f"Error calculating days from date: {e}")
        return 999


@router.get("/items")
async def get_audit_items(
    q: Optional[str] = None,
    category: Optional[str] = None,
    subcategory: Optional[str] = None,
    status_filter: Optional[str] = None,  # ✅ NEW: Filter by recount status
    sort_by: Optional[str] = 'category',
    limit: int = 100,  # За замовчуванням тільки 100 товарів
    db: Session = Depends(get_rh_db)  # ✅ MIGRATED: Using RentalHub DB
):
    """
    Отримати список товарів для переобліку
    ✅ MIGRATED: Using RentalHub DB (products + inventory tables)
    """
    try:
        # ✅ UPDATED: Using products + audit_records for status
        sql_parts = ["""
            SELECT 
                p.product_id,
                p.sku,
                p.name,
                p.category_name,
                p.subcategory_name,
                p.price,
                p.rental_price,
                p.image_url,
                p.description,
                p.care_instructions,
                p.color,
                p.material,
                p.size,
                p.quantity,
                p.zone,
                p.aisle,
                p.shelf,
                p.cleaning_status,
                p.product_state,
                p.last_audit_date,
                ar.status as audit_status,
                p.height_cm,
                p.width_cm,
                p.depth_cm,
                p.diameter_cm,
                p.shape,
                p.hashtags,
                p.status as product_status
            FROM products p
            LEFT JOIN (
                SELECT product_id, status 
                FROM audit_records ar1
                WHERE ar1.audit_date = (
                    SELECT MAX(ar2.audit_date) FROM audit_records ar2 WHERE ar2.product_id = ar1.product_id
                )
            ) ar ON p.product_id = ar.product_id
        """]
        
        # Default: show active only. 'disabled' filter overrides this.
        if status_filter == 'disabled':
            sql_parts.append("WHERE p.status = 0")
        else:
            sql_parts.append("WHERE p.status = 1")
        
        params = {}
        
        # Status filter — based on last_audit_date (same logic as stats)
        if status_filter and status_filter != 'all':
            if status_filter == 'critical':
                # Products with active damage records
                sql_parts.append("""AND p.product_id IN (
                    SELECT DISTINCT product_id FROM product_damage_history
                    WHERE COALESCE(processing_status, '') NOT IN ('completed', 'returned_to_stock', 'hidden', 'deleted')
                    AND processing_type IN ('wash', 'restoration', 'laundry')
                )""")
            elif status_filter == 'needs_recount' or status_filter == 'minor':
                sql_parts.append("AND (p.last_audit_date IS NULL OR DATEDIFF(CURDATE(), p.last_audit_date) > 180)")
            elif status_filter == 'ok':
                sql_parts.append("AND p.last_audit_date IS NOT NULL AND DATEDIFF(CURDATE(), p.last_audit_date) <= 180")
            elif status_filter == 'disabled':
                # Override the WHERE p.status = 1 — show disabled only
                pass  # handled below
        
        # Category filter
        if category and category != 'all':
            sql_parts.append("AND p.category_name = :category")
            params['category'] = category
        
        # Subcategory filter
        if subcategory and subcategory != 'all':
            sql_parts.append("AND p.subcategory_name = :subcategory")
            params['subcategory'] = subcategory
        
        # Search filter (SKU, name, category)
        if q:
            sql_parts.append("""
                AND (
                    p.sku LIKE :search
                    OR p.name LIKE :search
                    OR p.category_name LIKE :search
                    OR p.subcategory_name LIKE :search
                )
            """)
            params['search'] = f"%{q}%"
        
        # Sorting
        if sort_by == 'category':
            sql_parts.append("ORDER BY p.category_name, p.subcategory_name, p.name")
        elif sort_by == 'last_audit':
            sql_parts.append("ORDER BY p.last_audit_date DESC NULLS LAST")
        else:
            sql_parts.append("ORDER BY p.name")
        
        # Limit - якщо є фільтр по статусу, показуємо всі
        if status_filter and status_filter != 'all':
            sql_parts.append(f"LIMIT 1000")  # Без обмеження для фільтрованих
        else:
            sql_parts.append(f"LIMIT {limit}")
        
        final_sql = " ".join(sql_parts)
        results = db.execute(text(final_sql), params).fetchall()
        
        # ✅ NEW: Format results from RentalHub DB
        audit_items = []
        
        # Collect all product IDs first for batch lifecycle query
        all_product_ids = [row[0] for row in results]
        
        # Batch lifecycle metrics — one query for all products
        lifecycle_map = {}
        if all_product_ids:
            ids_str = ','.join(str(pid) for pid in all_product_ids)
            lifecycle_batch = db.execute(text(f"""
                SELECT oi.product_id,
                    COUNT(DISTINCT oi.order_id) as rentals_count,
                    COALESCE(SUM(oi.total_rental), 0) as total_profit
                FROM order_items oi
                WHERE oi.product_id IN ({ids_str})
                GROUP BY oi.product_id
            """)).fetchall()
            for lr in lifecycle_batch:
                lifecycle_map[lr[0]] = {'rentals': lr[1], 'profit': float(lr[2])}
            
            # Batch damages count
            damages_batch = db.execute(text(f"""
                SELECT product_id, COUNT(*) as cnt
                FROM product_damage_history
                WHERE product_id IN ({ids_str})
                GROUP BY product_id
            """)).fetchall()
            for dr in damages_batch:
                if dr[0] in lifecycle_map:
                    lifecycle_map[dr[0]]['damages'] = dr[1]
                else:
                    lifecycle_map[dr[0]] = {'rentals': 0, 'profit': 0, 'damages': dr[1]}
        
        for row in results:
            product_id = row[0]
            sku = row[1]
            name = row[2]
            category_name = row[3]
            subcategory_name = row[4]
            price = row[5]
            rental_price = row[6]
            image_url = row[7]
            description = row[8]
            care_instructions = row[9]
            color = row[10]
            material = row[11]
            size = row[12]
            quantity = row[13] or 0
            zone = row[14]
            aisle = row[15]
            shelf = row[16]
            cleaning_status = row[17]
            product_state = row[18]
            last_audit_date = row[19]
            audit_status_db = row[20]
            height_cm = row[21]
            width_cm = row[22]
            depth_cm = row[23]
            diameter_cm = row[24]
            shape = row[25]
            hashtags_json = row[26]
            product_status_val = row[27]
            
            import json
            try:
                hashtags = json.loads(hashtags_json) if hashtags_json else []
            except:
                hashtags = []
            
            cat_full = category_name or "Загальне"
            if subcategory_name:
                cat_full += f" · {subcategory_name}"
            
            last_audit = last_audit_date.isoformat() if last_audit_date else None
            days_from_audit = days_from_date(last_audit_date) if last_audit_date else 999
            
            # Lifecycle from batch query
            lc = lifecycle_map.get(product_id, {})
            total_rentals = lc.get('rentals', 0)
            total_profit = lc.get('profit', 0)
            damages_count = lc.get('damages', 0)
            
            photo_url = normalize_image_url(image_url)
            
            audit_items.append({
                'id': f"A-{product_id}",
                'product_id': product_id,
                'code': sku,
                'name': html.unescape(name) if name else name,
                'description': html.unescape(description) if description else description,
                'careInstructions': html.unescape(care_instructions) if care_instructions else None,
                'category': cat_full,
                'categoryName': category_name,
                'subcategoryName': subcategory_name,
                'zone': zone or '',
                'qty': quantity,
                'status': 'ok' if days_from_audit <= 180 else 'minor',
                'lastAuditDate': last_audit or '2024-01-01',
                'daysFromLastAudit': days_from_audit,
                'rentalsCount': total_rentals,
                'damagesCount': damages_count,
                'totalProfit': float(total_profit),
                'color': color,
                'material': material,
                'size': size,
                'heightCm': float(height_cm) if height_cm else None,
                'widthCm': float(width_cm) if width_cm else None,
                'depthCm': float(depth_cm) if depth_cm else None,
                'diameterCm': float(diameter_cm) if diameter_cm else None,
                'shape': shape,
                'hashtags': hashtags,
                'imageUrl': photo_url,
                'price': float(price) if price else 0,
                'rentalPrice': float(rental_price) if rental_price else 0,
                'cleaningStatus': cleaning_status or 'clean',
                'productState': product_state or 'available',
                'isDisabled': product_status_val == 0
            })
        
        return audit_items
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Помилка: {str(e)}")


@router.get("/items/{item_id}")
async def get_audit_item_details(
    item_id: str,
    db: Session = Depends(get_rh_db)
):
    """
    ✅ MIGRATED: Деталі товару для переобліку (RentalHub DB)
    """
    try:
        # Parse product_id from A-123
        product_id = int(item_id.replace('A-', ''))
        
        # ✅ UPDATED: Single query from products table (single source of truth)
        item_query = text("""
            SELECT 
                p.product_id, p.sku, p.name, p.description,
                p.category_name, p.subcategory_name,
                p.color, p.material, p.size, p.price, p.image_url,
                p.quantity, p.zone, p.aisle, p.shelf,
                p.cleaning_status, p.product_state,
                p.last_audit_date
            FROM products p
            WHERE p.product_id = :pid
        """)
        
        result = db.execute(item_query, {"pid": product_id}).fetchone()
        
        if not result:
            raise HTTPException(status_code=404, detail="Товар не знайдено")
        
        # Extract data (updated indices after removing inventory JOIN)
        quantity = result[11] or 0
        zone_raw = result[12] or ''
        aisle_raw = result[13] or ''
        shelf_raw = result[14] or ''
        
        # ✅ FIXED: Чисті значення + display значення
        zone_display = f"Зона {zone_raw}" if zone_raw else "Склад"
        location_parts = [aisle_raw, shelf_raw]
        location_display = " / ".join(filter(None, location_parts)) if any(location_parts) else "Не вказано"
        
        cat_full = result[4] or "Загальне"
        if result[5]:
            cat_full += f" · {result[5]}"
        
        last_audit = result[17].isoformat() if result[17] else None
        days_from_audit = days_from_date(result[17]) if result[17] else 999
        
        # ✅ ENHANCED: Lifecycle metrics from RentalHub DB
        
        # 1. Дані з OpenCart (старі)
        oc_rentals_query = text("""
            SELECT COUNT(DISTINCT op.order_id) as rentals_count,
                   COALESCE(SUM(op.total), 0) as total_profit
            FROM oc_order_product op
            INNER JOIN oc_order o ON op.order_id = o.order_id
            WHERE op.product_id = :product_id
            AND o.order_status_id NOT IN (7)
        """)
        oc_result = db.execute(oc_rentals_query, {"product_id": product_id}).fetchone()
        oc_rentals = oc_result[0] if oc_result else 0
        oc_profit = oc_result[1] if oc_result else 0
        
        # 2. Дані з нових таблиць
        from models_sqlalchemy import DecorOrderItem
        new_rentals = db.query(func.count(func.distinct(DecorOrderItem.order_id))).filter(
            DecorOrderItem.product_id == product_id
        ).scalar() or 0
        
        new_profit = db.query(func.sum(DecorOrderItem.total_rental)).filter(
            DecorOrderItem.product_id == product_id
        ).scalar() or 0
        
        # Об'єднати
        rentals_count = oc_rentals + new_rentals
        total_profit = float(oc_profit) + float(new_profit)
        
        # 3. Кейси шкоди
        damages_count = db.query(func.count(DecorDamageItem.id)).filter(
            DecorDamageItem.product_id == product_id
        ).scalar() or 0
        
        # 4. Останнє замовлення
        last_new_order = db.query(DecorOrderItem.order_id).filter(
            DecorOrderItem.product_id == product_id
        ).order_by(DecorOrderItem.created_at.desc()).first()
        
        if last_new_order:
            last_order_id = last_new_order[0]
        else:
            last_oc_order_query = text("""
                SELECT op.order_id
                FROM oc_order_product op
                WHERE op.product_id = :product_id
                ORDER BY op.order_product_id DESC
                LIMIT 1
            """)
            last_oc_result = db.execute(last_oc_order_query, {"product_id": product_id}).fetchone()
            last_order_id = last_oc_result[0] if last_oc_result else None
        
        # Отримати колір та матеріал з OpenCart атрибутів
        color_attr = db.execute(text("""
            SELECT text FROM oc_product_attribute 
            WHERE product_id = :pid AND attribute_id = 13 AND language_id = 4
        """), {"pid": product_id}).fetchone()
        
        material_attr = db.execute(text("""
            SELECT text FROM oc_product_attribute 
            WHERE product_id = :pid AND attribute_id = 16 AND language_id = 4
        """), {"pid": product_id}).fetchone()
        
        color_val = color_attr[0] if color_attr else None
        material_val = material_attr[0] if material_attr else None
        
        # Отримати розмір з extended
        size_val = extended.size if extended else None
        
        # Отримати фото
        image_url = normalize_image_url(product.image)
        
        # Отримати опис та інструкцію по догляду
        product_description = description.description if description and description.description else None
        
        # Отримати інструкцію з decor_product_extended (пріоритет)
        care_instructions = None
        if extended and extended.care_notes:
            care_instructions = extended.care_notes
        else:
            # Fallback: спробувати взяти з атрибутів OpenCart
            care_instructions_attr = db.execute(text("""
                SELECT text FROM oc_product_attribute 
                WHERE product_id = :pid AND attribute_id = 17 AND language_id = 4
            """), {"pid": product_id}).fetchone()
            care_instructions = care_instructions_attr[0] if care_instructions_attr else None
        
        return {
            'id': item_id,
            'product_id': product_id,
            'code': product.model,
            'name': description.name if description else 'Без назви',
            'description': product_description,
            'careInstructions': care_instructions,
            'category': cat_main,
            'zone': zone_raw or '',  # ✅ Єдине поле для локації
            'qty': int(product.quantity) if product.quantity else 0,
            'status': audit_status,
            'lastAuditDate': last_audit or '2024-01-01',
            'lastAuditBy': last_audit_by,
            'nextAuditDate': next_audit,
            'daysFromLastAudit': days_from_audit,
            'rentalsCount': rentals_count,
            'lastOrderId': last_order_id,
            'damagesCount': damages_count,
            'totalProfit': float(total_profit),
            'notes': audit_notes,
            'color': color_val,
            'material': material_val,
            'size': size_val,
            'imageUrl': image_url
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Помилка: {str(e)}"
        )


@router.put("/items/{item_id}/update-info")
async def update_product_info(
    item_id: str,
    data: dict,
    rh_db: Session = Depends(get_rh_db)
):
    """
    Оновити опис товару та інструкцію по догляду
    Зберігає ТІЛЬКИ в RentalHub БД
    """
    try:
        product_id = int(item_id.replace('A-', ''))
        
        # Оновити в RentalHub БД
        update_fields = []
        update_params = {'pid': product_id}
        
        if 'description' in data:
            update_fields.append("description = :description")
            update_params['description'] = data['description']
        
        if 'care_instructions' in data:
            update_fields.append("care_instructions = :care_instructions")
            update_params['care_instructions'] = data['care_instructions']
        
        if update_fields:
            update_sql = f"UPDATE products SET {', '.join(update_fields)} WHERE product_id = :pid"
            rh_db.execute(text(update_sql), update_params)
            rh_db.commit()
        
        return {
            'success': True,
            'message': 'Опис та інструкцію оновлено'
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Помилка оновлення: {str(e)}"
        )


@router.post("/items/{item_id}/mark-as-audited")
async def mark_audited_new_working_endpoint(item_id: str, audit_data: dict, db: Session = Depends(get_rh_db)):
    """
    ✅ MIGRATED: Фіксація переобліку в RentalHub DB
    Using: inventory + audit_records tables
    """
    try:
        product_id = int(item_id.replace('A-', ''))
        today = datetime.now().date()
        audited_by = audit_data.get('audited_by', 'Реквізитор')
        
        # Calculate next audit date
        next_audit_date = None
        if audit_data.get('next_audit_days'):
            days = int(audit_data['next_audit_days'])
            next_audit_date = (datetime.now() + timedelta(days=days)).date()
        
        # ✅ UPDATED: Update products table (single source of truth) in RentalHub DB
        # Get current quantity
        current_qty_result = db.execute(text("""
            SELECT quantity FROM products WHERE product_id = :pid
        """), {"pid": product_id})
        current_qty_row = current_qty_result.fetchone()
        quantity_expected = current_qty_row[0] if current_qty_row else 0
        
        # Get actual quantity from audit data
        quantity_actual = audit_data.get('quantity_actual', quantity_expected)
        
        update_products = text("""
            UPDATE products 
            SET last_audit_date = :last_audit_date,
                quantity = :quantity_actual
            WHERE product_id = :product_id
        """)
        
        result = db.execute(update_products, {
            'product_id': product_id,
            'last_audit_date': today,
            'quantity_actual': quantity_actual
        })
        
        # ✅ UPDATED: Insert audit record with quantity tracking
        audit_id = str(uuid.uuid4())
        audit_insert = text("""
            INSERT INTO audit_records 
            (id, product_id, audit_date, audited_by, status, quantity_expected, quantity_actual, notes)
            VALUES 
            (:audit_id, :product_id, CURDATE(), :audited_by, :audit_status, :qty_expected, :qty_actual, :notes)
        """)
        
        db.execute(audit_insert, {
            'audit_id': audit_id,
            'product_id': product_id,
            'audited_by': audited_by,
            'audit_status': audit_data.get('audit_status', 'ok'),
            'qty_expected': quantity_expected,
            'qty_actual': quantity_actual,
            'notes': audit_data.get('notes', '')
        })
        
        db.commit()
        
        return {
            'success': True,
            'message': 'Переоблік зафіксовано в RentalHub DB',
            'item_id': item_id,
            'audit_id': audit_id,
            'rows_updated': result.rowcount
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Помилка: {str(e)}")

@router.put("/items/{item_id}/audit-DISABLED-FOR-TEST")
async def mark_item_audited_v2_fixed(
    item_id: str,
    audit_data: dict,
    db: Session = Depends(get_rh_db)
):
    """
    DISABLED FOR TESTING
    """
    import sys
    sys.stdout.flush()
    sys.stderr.write(f"[AUDIT ENDPOINT] ✅ ВИКЛИКАНО! item_id={item_id}\n")
    sys.stderr.write(f"[AUDIT ENDPOINT] audit_data={audit_data}\n")
    sys.stderr.flush()
    
    try:
        product_id = int(item_id.replace('A-', ''))
        
        sys.stderr.write(f"[AUDIT] product_id={product_id}\n")
        sys.stderr.write(f"[AUDIT] Готуємо оновлення для DecorProductCatalog...\n")
        sys.stderr.flush()
        
        # Оновити DecorProductCatalog через raw SQL (більш надійно)
        today = datetime.now().date()
        audited_by = audit_data.get('audited_by', 'Реквізитор')
        
        sys.stderr.write(f"[AUDIT] 📝 Оновлюємо catalog для product_id={product_id}\n")
        sys.stderr.write(f"[AUDIT] 📅 Дата: {today}, Хто: {audited_by}\n")
        sys.stderr.flush()
        
        # Спробувати INSERT або UPDATE
        update_catalog = text("""
            INSERT INTO decor_product_catalog 
            (product_id, last_audit_date, last_audit_by, next_audit_date, created_at)
            VALUES 
            (:product_id, :last_audit_date, :last_audit_by, :next_audit_date, NOW())
            ON DUPLICATE KEY UPDATE
            last_audit_date = :last_audit_date,
            last_audit_by = :last_audit_by,
            next_audit_date = :next_audit_date
        """)
        
        next_audit_date = None
        if audit_data.get('next_audit_days'):
            days = int(audit_data['next_audit_days'])
            next_audit_date = (datetime.now() + timedelta(days=days)).date()
            sys.stderr.write(f"[AUDIT] Setting next_audit_date={next_audit_date} (+{days} days)\n")
            sys.stderr.flush()
        
        sys.stderr.write(f"[AUDIT] Виконуємо SQL UPDATE...\n")
        sys.stderr.flush()
        
        db.execute(update_catalog, {
            'product_id': product_id,
            'last_audit_date': today,
            'last_audit_by': audited_by,
            'next_audit_date': next_audit_date
        })
        
        # Створити запис в історії
        history = DecorProductHistory(
            id=f"H-{uuid.uuid4().hex[:8].upper()}",
            product_id=product_id,
            inventory_item_id=None,
            event_type='cleaned',  # audit event
            actor=audit_data.get('audited_by', 'Реквізитор'),
            notes=f"Переоблік: {audit_data.get('notes', 'стан перевірено')}"
        )
        db.add(history)
        
        # Створити запис в decor_product_audits
        audit_id = f"AUDIT-{uuid.uuid4().hex[:8].upper()}"
        audit_insert = text("""
            INSERT INTO decor_product_audits 
            (id, product_id, audit_date, audited_by, audit_status, audit_notes)
            VALUES 
            (:audit_id, :product_id, CURDATE(), :audited_by, :audit_status, :notes)
        """)
        db.execute(audit_insert, {
            'audit_id': audit_id,
            'product_id': product_id,
            'audited_by': audit_data.get('audited_by', 'Реквізитор'),
            'audit_status': audit_data.get('audit_status', 'ok'),
            'notes': audit_data.get('notes', '')
        })
        
        # Явно commit всі зміни
        sys.stderr.write(f"[AUDIT] Викликаємо db.commit()...\n")
        sys.stderr.flush()
        
        db.commit()
        
        sys.stderr.write(f"[AUDIT] ✅ Committed successfully!\n")
        sys.stderr.flush()
        
        return {
            'success': True,
            'message': 'ТЕСТОВЕ ПОВІДОМЛЕННЯ ВІД НОВОГО КОДУ 2025-11-13',
            'item_id': item_id,
            'audit_id': audit_id,
            'debug': 'Код точно виконується!'
        }
        
    except Exception as e:
        sys.stderr.write(f"[AUDIT] ❌ Error: {str(e)}\n")
        sys.stderr.flush()
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Помилка: {str(e)}"
        )


@router.get("/stats")
async def get_audit_stats(
    db: Session = Depends(get_rh_db)
):
    """
    ✅ Статистика по переобліку з оновленою логікою:
    - ok (Задовільний стан) = переоблікавані товари (last_audit_date <= 180 днів)
    - minor (Потребують уваги) = не переоблікавані товари (немає дати або > 180 днів)
    - crit (Критичні) = товари що є в кабінеті шкоди
    """
    try:
        # Підрахунок по переобліку
        stats_query = text("""
            SELECT 
                COUNT(*) as total,
                -- Задовільний стан: переоблікавані (дата є і <= 180 днів)
                SUM(CASE 
                    WHEN p.last_audit_date IS NOT NULL 
                    AND DATEDIFF(CURDATE(), p.last_audit_date) <= 180 
                    THEN 1 ELSE 0 
                END) as ok,
                -- Потребують уваги: не переоблікавані (немає дати або > 180 днів)
                SUM(CASE 
                    WHEN p.last_audit_date IS NULL 
                    OR DATEDIFF(CURDATE(), p.last_audit_date) > 180 
                    THEN 1 ELSE 0 
                END) as minor
            FROM products p
            WHERE p.status = 1
        """)
        
        result = db.execute(stats_query).fetchone()
        
        # Підрахунок критичних (активні записи в product_damage_history)
        damages_query = text("""
            SELECT COUNT(*) as damaged
            FROM product_damage_history
            WHERE COALESCE(processing_status, '') NOT IN ('completed', 'returned_to_stock', 'hidden', 'deleted')
            AND processing_type IN ('wash', 'restoration', 'laundry')
        """)
        
        damages_result = db.execute(damages_query).fetchone()
        
        stats = {
            'total': result[0] or 0,
            'ok': result[1] or 0,        # Переоблікавані
            'minor': result[2] or 0,     # Не переоблікавані
            'crit': damages_result[0] or 0,  # Критичні (шкоди)
            'lost': 0,  # Deprecated
            'overdueCnt': result[2] or 0  # Backwards compatibility
        }
        
        return stats
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Помилка: {str(e)}"
        )


@router.post("/calculate-lifecycle/{product_id}")
async def calculate_lifecycle_metrics(
    product_id: int,
    db: Session = Depends(get_rh_db)
):
    """
    Підрахувати lifecycle метрики для товару
    """
    try:
        from models_sqlalchemy import DecorProductLifecycle, DecorOrderItem, FinanceTransaction
        
        # 1. Рахуємо кількість оренд через DecorIssueCard
        rentals_count = db.query(func.count(DecorIssueCard.id)).filter(
            DecorIssueCard.status.in_(['issued', 'archived'])
        ).join(
            DecorOrder,
            DecorIssueCard.order_id == DecorOrder.id
        ).join(
            DecorOrderItem,
            DecorOrderItem.order_id == DecorOrder.id
        ).filter(
            DecorOrderItem.product_id == product_id
        ).scalar() or 0
        
        # 2. Рахуємо кількість пошкоджень з decor_damages
        damages_count = db.query(func.count(DecorDamage.id)).join(
            DecorDamageItem,
            DecorDamageItem.damage_id == DecorDamage.id
        ).filter(
            DecorDamageItem.product_id == product_id
        ).scalar() or 0
        
        # 3. Рахуємо загальний прибуток з цього товару
        # Сумуємо rental суми з DecorOrderItem де product_id відповідає
        total_profit = db.query(func.sum(DecorOrderItem.total_rental)).filter(
            DecorOrderItem.product_id == product_id
        ).scalar() or 0
        
        # 4. Знаходимо останнє замовлення з цим товаром
        last_order_item = db.query(DecorOrderItem).filter(
            DecorOrderItem.product_id == product_id
        ).order_by(DecorOrderItem.created_at.desc()).first()
        
        last_order_id = last_order_item.order_id if last_order_item else None
        
        # 5. Зберігаємо/оновлюємо в decor_product_lifecycle
        lifecycle = db.query(DecorProductLifecycle).filter(
            DecorProductLifecycle.product_id == product_id
        ).first()
        
        if lifecycle:
            # Оновлюємо існуючий запис
            lifecycle.rentals_count = rentals_count
            lifecycle.damages_count = damages_count
            lifecycle.total_profit = total_profit
            lifecycle.last_rental_order_id = last_order_id
            lifecycle.last_updated = datetime.now()
        else:
            # Створюємо новий запис
            lifecycle = DecorProductLifecycle(
                product_id=product_id,
                rentals_count=rentals_count,
                damages_count=damages_count,
                total_profit=total_profit,
                last_rental_order_id=last_order_id
            )
            db.add(lifecycle)
        
        db.commit()
        
        return {
            'success': True,
            'product_id': product_id,
            'rentals_count': rentals_count,
            'damages_count': damages_count,
            'total_profit': float(total_profit),
            'last_order_id': last_order_id
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Помилка підрахунку: {str(e)}"
        )


@router.put("/items/{item_id}/quantity")
async def update_quantity(
    item_id: str,
    data: dict,
    db: Session = Depends(get_rh_db)
):
    """
    Оновити кількість товару
    ✅ MIGRATED: Updates products.quantity in RentalHub DB (single source of truth)
    """
    try:
        product_id = int(item_id.replace('A-', ''))
        new_qty = data.get('qty')
        
        if new_qty is None or new_qty < 0:
            raise HTTPException(status_code=400, detail="Невірна кількість")
        
        # Get current quantity from RentalHub DB
        result = db.execute(text("""
            SELECT quantity FROM products WHERE product_id = :id
        """), {"id": product_id})
        
        row = result.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Товар не знайдено")
        
        old_qty = row[0] or 0
        
        # Update quantity in RentalHub DB (single source of truth)
        db.execute(text("""
            UPDATE products SET quantity = :qty WHERE product_id = :id
        """), {"qty": new_qty, "id": product_id})
        
        db.commit()
        
        return {
            'success': True,
            'product_id': product_id,
            'old_qty': old_qty,
            'new_qty': new_qty,
            'message': 'Кількість оновлена в RentalHub DB (джерело правди)'
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Помилка: {str(e)}")


@router.put("/items/{item_id}/location")
async def update_location(
    item_id: str,
    data: dict,
    db: Session = Depends(get_rh_db)
):
    """
    Оновити локацію товару
    ✅ MIGRATED: Updates products table (single source of truth)
    """
    try:
        product_id = int(item_id.replace('A-', ''))
        
        zone = data.get('zone')
        location = data.get('location')
        aisle = None
        shelf = None
        
        if location:
            # Parse location (напр. "6A12" -> aisle="6A", shelf="12")
            aisle = location[:2] if len(location) >= 2 else location
            shelf = location[2:] if len(location) > 2 else None
        
        # ✅ UPDATED: Update products table (single source of truth)
        db.execute(text("""
            UPDATE products 
            SET zone = :zone, aisle = :aisle, shelf = :shelf
            WHERE product_id = :pid
        """), {"zone": zone, "aisle": aisle, "shelf": shelf, "pid": product_id})
        
        db.commit()
        
        return {
            'success': True,
            'product_id': product_id,
            'zone': zone,
            'location': location,
            'message': 'Локацію оновлено в products (джерело правди)'
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Помилка: {str(e)}")


@router.put("/items/{item_id}/status")
async def update_audit_status(
    item_id: str,
    data: dict,
    db: Session = Depends(get_rh_db)
):
    """
    Оновити audit status товару
    """
    try:
        product_id = int(item_id.replace('A-', ''))
        new_status = data.get('status')
        
        if new_status not in ['ok', 'minor', 'critical', 'lost']:
            raise HTTPException(status_code=400, detail="Невірний статус")
        
        # Оновити в inventory
        inventory = db.query(DecorInventoryItem).filter(
            DecorInventoryItem.product_id == product_id
        ).first()
        
        if not inventory:
            inventory = DecorInventoryItem(
                id=f"INV-{uuid.uuid4().hex[:8].upper()}",
                product_id=product_id,
                inventory_code=f"AUTO-{product_id}",
                status='available'
            )
            db.add(inventory)
        
        old_status = inventory.audit_status
        inventory.audit_status = new_status
        
        # Історія
        history = DecorProductHistory(
            id=f"H-{uuid.uuid4().hex[:8].upper()}",
            product_id=product_id,
            event_type='edited',
            actor=data.get('actor', 'Реквізитор'),
            notes=f"Статус змінено з {old_status} на {new_status}"
        )
        db.add(history)
        
        db.commit()
        
        return {
            'success': True,
            'product_id': product_id,
            'old_status': old_status,
            'new_status': new_status
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Помилка: {str(e)}")


@router.put("/items/{item_id}/notes")
async def update_notes(
    item_id: str,
    data: dict,
    db: Session = Depends(get_rh_db)
):
    """
    Оновити нотатки
    """
    try:
        product_id = int(item_id.replace('A-', ''))
        notes = data.get('notes', '')
        
        inventory = db.query(DecorInventoryItem).filter(
            DecorInventoryItem.product_id == product_id
        ).first()
        
        if not inventory:
            inventory = DecorInventoryItem(
                id=f"INV-{uuid.uuid4().hex[:8].upper()}",
                product_id=product_id,
                inventory_code=f"AUTO-{product_id}",
                status='available'
            )
            db.add(inventory)
        
        inventory.audit_notes = notes
        
        db.commit()
        
        return {
            'success': True,
            'product_id': product_id
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Помилка: {str(e)}")


@router.get("/items/{item_id}/history")
async def get_product_history(
    item_id: str,
    db: Session = Depends(get_rh_db)
):
    """
    Отримати історію товару
    """
    try:
        product_id = int(item_id.replace('A-', ''))
        
        history_items = db.query(DecorProductHistory).filter(
            DecorProductHistory.product_id == product_id
        ).order_by(DecorProductHistory.created_at.desc()).limit(50).all()
        
        history_list = []
        for h in history_items:
            history_list.append({
                'id': h.id,
                'date': h.created_at.isoformat() if h.created_at else None,
                'kind': h.event_type,
                'actor': h.actor or 'Система',
                'orderId': h.order_id,
                'note': h.notes or ''
            })
        
        return history_list
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Помилка: {str(e)}")


@router.get("/items/{item_id}/rental-history")
async def get_rental_history(
    item_id: str,
    db: Session = Depends(get_rh_db)
):
    """
    Отримати історію оренд товару з RentalHub (orders + order_items)
    """
    try:
        product_id = int(item_id.replace('A-', ''))
        
        query = text("""
            SELECT 
                o.order_id,
                o.order_number,
                o.customer_name as client_name,
                COALESCE(o.customer_phone, o.phone) as client_phone,
                o.rental_start_date as rent_date,
                o.rental_end_date as rent_return_date,
                o.rental_days,
                oi.quantity,
                oi.total_rental,
                o.status,
                o.created_at
            FROM order_items oi
            INNER JOIN orders o ON oi.order_id = o.order_id
            WHERE oi.product_id = :product_id
            ORDER BY o.created_at DESC
            LIMIT 50
        """)
        
        result = db.execute(query, {"product_id": product_id})
        rows = result.fetchall()
        
        rental_history = []
        for row in rows:
            rent_date = row[4].isoformat() if hasattr(row[4], 'isoformat') else str(row[4]) if row[4] else None
            rent_return_date = row[5].isoformat() if hasattr(row[5], 'isoformat') else str(row[5]) if row[5] else None
            created_at = row[10].isoformat() if hasattr(row[10], 'isoformat') else str(row[10]) if row[10] else None
            
            rental_history.append({
                'order_id': row[0],
                'order_number': f"#{row[1]}" if row[1] else f"#{row[0]}",
                'client_name': row[2] or 'Клієнт не вказаний',
                'client_phone': row[3] or '',
                'rent_date': rent_date,
                'rent_return_date': rent_return_date,
                'rental_days': row[6] or 1,
                'quantity': row[7],
                'total_rental': float(row[8]) if row[8] else 0.0,
                'deposit': 0.0,
                'status': row[9] or 'Невідомо',
                'created_at': created_at
            })
        
        return rental_history
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Помилка: {str(e)}")


@router.get("/items/{item_id}/audit-history")
async def get_audit_history(
    item_id: str,
    db: Session = Depends(get_rh_db)
):
    """
    Отримати історію переобліків товару
    """
    try:
        product_id = int(item_id.replace('A-', ''))
        
        query = text("""
            SELECT 
                id,
                audit_date,
                audited_by,
                audit_status,
                audit_notes,
                created_at
            FROM decor_product_audits
            WHERE product_id = :product_id
            ORDER BY audit_date DESC, created_at DESC
            LIMIT 50
        """)
        
        result = db.execute(query, {"product_id": product_id})
        rows = result.fetchall()
        
        audit_history = []
        for row in rows:
            audit_date = row[1].isoformat() if hasattr(row[1], 'isoformat') else str(row[1])
            created_at = row[5].isoformat() if hasattr(row[5], 'isoformat') else str(row[5])
            
            audit_history.append({
                'id': row[0],
                'audit_date': audit_date,
                'audited_by': row[2] or 'Невідомо',
                'audit_status': row[3] or 'ok',
                'audit_notes': row[4] or '',
                'created_at': created_at
            })
        
        return audit_history
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Помилка: {str(e)}")


# ============================================================
# ACTION BUTTON ENDPOINTS



@router.put("/items/{item_id}/edit-full")
async def edit_item_full(
    item_id: str,
    data: dict,
    rh_db: Session = Depends(get_rh_db)
):
    """
    Повне редагування товару: назва, колір, матеріал, кількість, розташування,
    категорія, підкатегорія, розміри (окремо), форма, хештеги
    Зберігає ТІЛЬКИ в RentalHub БД
    """
    try:
        product_id = int(item_id.replace('A-', ''))
        changes = []
        
        # Оновити в RentalHub БД
        update_fields = []
        update_params = {'pid': product_id}
        
        if 'name' in data and data['name']:
            update_fields.append("name = :name")
            update_params['name'] = data['name']
            changes.append(f"Назва: {data['name']}")
        
        if 'code' in data and data['code']:
            update_fields.append("sku = :sku")
            update_params['sku'] = data['code']
            changes.append(f"Код: {data['code']}")
        
        if 'price' in data:
            update_fields.append("price = :price")
            update_params['price'] = data['price']
            changes.append(f"Ціна купівлі: ₴{data['price']}")
        
        if 'rentalPrice' in data:
            update_fields.append("rental_price = :rental_price")
            update_params['rental_price'] = data['rentalPrice']
            changes.append(f"Ціна оренди: ₴{data['rentalPrice']}/день")
        
        if 'color' in data:
            update_fields.append("color = :color")
            update_params['color'] = data['color']
            changes.append(f"Колір: {data['color']}")
        
        if 'material' in data:
            update_fields.append("material = :material")
            update_params['material'] = data['material']
            changes.append(f"Матеріал: {data['material']}")
        
        # === НОВІ ПОЛЯ: Розміри окремо ===
        if 'height' in data:
            update_fields.append("height_cm = :height")
            update_params['height'] = float(data['height']) if data['height'] else None
            if data['height']:
                changes.append(f"Висота: {data['height']} см")
        
        if 'width' in data:
            update_fields.append("width_cm = :width")
            update_params['width'] = float(data['width']) if data['width'] else None
            if data['width']:
                changes.append(f"Ширина: {data['width']} см")
        
        if 'depth' in data:
            update_fields.append("depth_cm = :depth")
            update_params['depth'] = float(data['depth']) if data['depth'] else None
            if data['depth']:
                changes.append(f"Глибина: {data['depth']} см")
        
        if 'diameter' in data:
            update_fields.append("diameter_cm = :diameter")
            update_params['diameter'] = float(data['diameter']) if data['diameter'] else None
            if data['diameter']:
                changes.append(f"Діаметр: {data['diameter']} см")
        
        # Також оновимо старе поле size для сумісності
        if any(k in data for k in ['height', 'width', 'depth']):
            h = data.get('height', '')
            w = data.get('width', '')
            d = data.get('depth', '')
            size_parts = [str(x) for x in [h, w, d] if x]
            if size_parts:
                size_str = 'x'.join(size_parts)
                update_fields.append("size = :size")
                update_params['size'] = size_str
        
        # === НОВІ ПОЛЯ: Форма ===
        if 'shape' in data:
            update_fields.append("shape = :shape")
            update_params['shape'] = data['shape']
            if data['shape']:
                changes.append(f"Форма: {data['shape']}")
        
        # === НОВІ ПОЛЯ: Категорія/Підкатегорія ===
        if 'category' in data:
            update_fields.append("category_name = :category")
            update_params['category'] = data['category']
            if data['category']:
                changes.append(f"Категорія: {data['category']}")
        
        if 'subcategory' in data:
            update_fields.append("subcategory_name = :subcategory")
            update_params['subcategory'] = data['subcategory']
            if data['subcategory']:
                changes.append(f"Підкатегорія: {data['subcategory']}")
        
        # === НОВІ ПОЛЯ: Хештеги (JSON array) ===
        if 'hashtags' in data:
            import json
            hashtags = data['hashtags'] if isinstance(data['hashtags'], list) else []
            update_fields.append("hashtags = :hashtags")
            update_params['hashtags'] = json.dumps(hashtags, ensure_ascii=False)
            if hashtags:
                changes.append(f"Хештеги: {', '.join(['#' + t for t in hashtags])}")
        
        if 'qty' in data and data['qty'] is not None:
            update_fields.append("quantity = :qty")
            update_params['qty'] = data['qty']
            changes.append(f"Кількість: {data['qty']}")
        
        if 'zone' in data:
            update_fields.append("zone = :zone")
            update_params['zone'] = data['zone']
            changes.append(f"Зона: {data['zone']}")
        
        if 'location' in data and data['location']:
            # Парсинг локації (напр. "A1-10" -> aisle=A1, shelf=10)
            import re
            location = data['location']
            match = re.match(r'([A-Z]\d+)-(\d+)', location)
            if match:
                update_fields.append("aisle = :aisle")
                update_fields.append("shelf = :shelf")
                update_params['aisle'] = match.group(1)
                update_params['shelf'] = match.group(2)
            else:
                update_fields.append("aisle = :aisle")
                update_params['aisle'] = location
            changes.append(f"Локація: {location}")
        
        # Виконати оновлення
        if update_fields:
            update_sql = f"UPDATE products SET {', '.join(update_fields)} WHERE product_id = :pid"
            rh_db.execute(text(update_sql), update_params)
            rh_db.commit()
        
        return {
            'success': True,
            'product_id': product_id,
            'changes': changes,
            'message': 'Дані оновлено успішно'
        }
        
    except Exception as e:
        rh_db.rollback()
        raise HTTPException(status_code=500, detail=f"Помилка: {str(e)}")



# ============================================================
# HASHTAGS DICTIONARY API
# ============================================================

@router.get("/hashtags")
async def list_hashtags(
    category: Optional[str] = None,
    rh_db: Session = Depends(get_rh_db)
):
    """
    Отримати список всіх хештегів зі словника
    """
    try:
        sql = """
            SELECT id, tag, display_name, category, usage_count, is_active
            FROM product_hashtags_dict
            WHERE is_active = TRUE
        """
        params = {}
        
        if category:
            sql += " AND category = :category"
            params['category'] = category
        
        sql += " ORDER BY category, usage_count DESC, tag"
        
        rows = rh_db.execute(text(sql), params).fetchall()
        
        return {
            "hashtags": [
                {
                    "id": r[0],
                    "tag": r[1],
                    "display_name": r[2],
                    "category": r[3],
                    "usage_count": r[4],
                    "is_active": r[5]
                }
                for r in rows
            ]
        }
    except Exception as e:
        return {"hashtags": [], "error": str(e)}


@router.post("/hashtags")
async def create_hashtag(
    data: dict,
    rh_db: Session = Depends(get_rh_db)
):
    """
    Створити новий хештег в словнику
    """
    try:
        tag = data.get('tag', '').lower().strip().replace(' ', '_').replace('#', '')
        display_name = data.get('display_name', tag.replace('_', ' ').title())
        category = data.get('category', 'general')
        
        if not tag:
            raise HTTPException(status_code=400, detail="Хештег не може бути пустим")
        
        # Перевірити чи існує
        existing = rh_db.execute(text("""
            SELECT id FROM product_hashtags_dict WHERE tag = :tag
        """), {"tag": tag}).fetchone()
        
        if existing:
            return {"success": False, "error": "Такий хештег вже існує", "id": existing[0]}
        
        rh_db.execute(text("""
            INSERT INTO product_hashtags_dict (tag, display_name, category)
            VALUES (:tag, :display_name, :category)
        """), {"tag": tag, "display_name": display_name, "category": category})
        
        new_id = rh_db.execute(text("SELECT LAST_INSERT_ID()")).fetchone()[0]
        rh_db.commit()
        
        return {
            "success": True,
            "id": new_id,
            "tag": tag,
            "display_name": display_name,
            "category": category
        }
    except HTTPException:
        raise
    except Exception as e:
        rh_db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/hashtags/categories")
async def list_hashtag_categories(rh_db: Session = Depends(get_rh_db)):
    """
    Отримати унікальні категорії хештегів
    """
    try:
        rows = rh_db.execute(text("""
            SELECT DISTINCT category, COUNT(*) as count
            FROM product_hashtags_dict
            WHERE is_active = TRUE
            GROUP BY category
            ORDER BY count DESC
        """)).fetchall()
        
        return {
            "categories": [
                {"code": r[0], "count": r[1]}
                for r in rows
            ]
        }
    except Exception as e:
        return {"categories": [], "error": str(e)}


@router.get("/shapes")
async def list_shapes(rh_db: Session = Depends(get_rh_db)):
    """
    Отримати унікальні форми виробів (для автозаповнення)
    """
    try:
        rows = rh_db.execute(text("""
            SELECT DISTINCT shape, COUNT(*) as count
            FROM products
            WHERE shape IS NOT NULL AND shape != ''
            GROUP BY shape
            ORDER BY count DESC
        """)).fetchall()
        
        # Додамо базові форми якщо список порожній
        base_shapes = [
            "круглий", "квадратний", "прямокутний", "овальний",
            "конусний", "циліндричний", "нестандартний"
        ]
        
        shapes = [{"shape": r[0], "count": r[1]} for r in rows]
        
        if not shapes:
            shapes = [{"shape": s, "count": 0} for s in base_shapes]
        
        return {"shapes": shapes}
    except Exception as e:
        return {"shapes": [], "error": str(e)}


@router.get("/categories-list")
async def list_categories_for_edit(rh_db: Session = Depends(get_rh_db)):
    """
    Отримати список категорій та підкатегорій для редагування товару
    """
    try:
        # Категорії
        cat_rows = rh_db.execute(text("""
            SELECT DISTINCT category_name, COUNT(*) as count
            FROM products
            WHERE category_name IS NOT NULL AND category_name != ''
            GROUP BY category_name
            ORDER BY count DESC
        """)).fetchall()
        
        # Підкатегорії згруповані по категоріях
        subcat_rows = rh_db.execute(text("""
            SELECT category_name, subcategory_name, COUNT(*) as count
            FROM products
            WHERE subcategory_name IS NOT NULL AND subcategory_name != ''
            GROUP BY category_name, subcategory_name
            ORDER BY category_name, count DESC
        """)).fetchall()
        
        # Групуємо підкатегорії по категоріях
        subcategories_by_cat = {}
        for r in subcat_rows:
            cat = r[0] or "Інше"
            if cat not in subcategories_by_cat:
                subcategories_by_cat[cat] = []
            subcategories_by_cat[cat].append({"name": r[1], "count": r[2]})
        
        return {
            "categories": [{"name": r[0], "count": r[1]} for r in cat_rows],
            "subcategories_by_category": subcategories_by_cat
        }
    except Exception as e:
        return {"categories": [], "subcategories_by_category": {}, "error": str(e)}


# ============================================================

@router.post("/items/{item_id}/send-to-wash")
async def send_to_wash(
    item_id: str,
    data: dict,
    db: Session = Depends(get_rh_db)
):
    """
    Відправити товар на мийку
    """
    try:
        product_id = int(item_id.replace('A-', ''))
        
        # Оновити статус в каталозі
        catalog = db.query(DecorProductCatalog).filter(
            DecorProductCatalog.product_id == product_id
        ).first()
        
        if not catalog:
            catalog = DecorProductCatalog(product_id=product_id)
            db.add(catalog)
        
        catalog.cleaning_status = 'wash'
        catalog.cleaning_last_updated = datetime.now()
        
        # Оновити inventory status
        inventory = db.query(DecorInventoryItem).filter(
            DecorInventoryItem.product_id == product_id
        ).first()
        
        if inventory:
            inventory.status = 'washing'
        
        # Історія
        history = DecorProductHistory(
            id=f"H-{uuid.uuid4().hex[:8].upper()}",
            product_id=product_id,
            event_type='cleaning',
            actor=data.get('actor', 'Реквізитор'),
            notes=f"Відправлено на мийку: {data.get('notes', '')}"
        )
        db.add(history)
        
        db.commit()
        
        return {
            'success': True,
            'product_id': product_id,
            'message': 'Товар відправлено на мийку'
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Помилка: {str(e)}")


@router.post("/items/{item_id}/add-damage")
async def add_damage_during_audit(
    item_id: str,
    data: dict,
    db: Session = Depends(get_rh_db)
):
    """
    Зафіксувати пошкодження при переобліку
    """
    try:
        product_id = int(item_id.replace('A-', ''))
        
        # Отримати інформацію про товар
        product = db.query(OpenCartProduct).filter(
            OpenCartProduct.product_id == product_id
        ).first()
        
        if not product:
            raise HTTPException(status_code=404, detail="Товар не знайдено")
        
        description = db.query(OpenCartProductDescription).filter(
            OpenCartProductDescription.product_id == product_id,
            OpenCartProductDescription.language_id == 4
        ).first()
        
        # Знайти останнє замовлення (якщо є)
        inventory = db.query(DecorInventoryItem).filter(
            DecorInventoryItem.product_id == product_id
        ).first()
        
        order_id = inventory.last_order_id if inventory and inventory.last_order_id else None
        
        # Створити запис пошкодження в історії
        damage_type = data.get('damage_type', 'physical')
        damage_description = data.get('description', '')
        severity = data.get('severity', 'minor')  # minor, critical
        photo_url = data.get('photo_url', '')
        
        # Записати в історію товару з фото
        history_id = f"H-{uuid.uuid4().hex[:8].upper()}"
        
        db.execute(text("""
            INSERT INTO product_damage_history 
            (id, product_id, created_by, order_id, note, photo_url, created_at)
            VALUES (:id, :product_id, :created_by, :order_id, :note, :photo_url, NOW())
        """), {
            'id': history_id,
            'product_id': product_id,
            'created_by': data.get('actor', 'Реквізитор'),
            'order_id': order_id,
            'note': f"Пошкодження ({severity}): {damage_description}",
            'photo_url': photo_url
        })
        
        # Оновити статус inventory якщо пошкодження критичне
        if inventory:
            if severity == 'critical':
                inventory.audit_status = 'critical'
            elif severity == 'minor' and inventory.audit_status != 'critical':
                inventory.audit_status = 'minor'
            
            # Додати нотатку про пошкодження
            current_notes = inventory.audit_notes or ''
            damage_note = f"\n[{datetime.now().strftime('%Y-%m-%d')}] Пошкодження: {damage_description}"
            inventory.audit_notes = (current_notes + damage_note)[:500]  # Обмеження 500 символів
        
        # Якщо потрібно створити повний кейс пошкодження
        if data.get('create_damage_case', False):
            damage_id = f"DMG-{uuid.uuid4().hex[:8].upper()}"
            damage = DecorDamage(
                id=damage_id,
                order_id=order_id,
                order_number=f"AUDIT-{product_id}",
                customer_id=0,
                customer_name="Переоблік",
                case_status='draft',
                finance_status='none',
                fulfillment_status='none',
                claimed_total=float(data.get('estimated_cost', 0)),
                notes=damage_description,
                created_by=data.get('actor', 'Реквізитор')
            )
            db.add(damage)
            
            # Додати item в кейс
            damage_item = DecorDamageItem(
                damage_id=damage_id,
                product_id=product_id,
                barcode=product.model,
                name=description.name if description else product.model,
                image=product.image,
                damage_type=damage_type,
                qty=1,
                base_value=float(product.ean) if product.ean else 0,
                estimate_value=float(data.get('estimated_cost', 0)),
                resolution='pending',
                comment=damage_description
            )
            db.add(damage_item)
        
        db.commit()
        
        return {
            'success': True,
            'product_id': product_id,
            'message': 'Пошкодження зафіксовано',
            'severity': severity
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Помилка: {str(e)}")


@router.get("/items/{item_id}/damages")
async def get_item_damages(
    item_id: str,
    db: Session = Depends(get_rh_db)
):
    """
    Отримати історію пошкоджень товару
    """
    try:
        product_id = int(item_id.replace('A-', ''))
        
        # Отримати всі записи про пошкодження з історії включаючи фото
        result = db.execute(text("""
            SELECT id, created_at, created_by as actor, note as notes, order_id, photo_url
            FROM product_damage_history
            WHERE product_id = :pid 
            ORDER BY created_at DESC
        """), {'pid': product_id})
        
        damage_list = []
        for row in result:
            damage_list.append({
                'id': row[0],
                'date': row[1].isoformat() if row[1] else None,
                'actor': row[2] or 'Система',
                'notes': row[3],
                'order_id': row[4],
                'photo_url': row[5]
            })
        
        return damage_list
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Помилка: {str(e)}")


@router.post("/items/{item_id}/send-to-restoration")
async def send_to_restoration(
    item_id: str,
    data: dict,
    db: Session = Depends(get_rh_db)
):
    """
    Відправити товар на реставрацію
    """
    try:
        product_id = int(item_id.replace('A-', ''))
        
        # Оновити статус в каталозі
        catalog = db.query(DecorProductCatalog).filter(
            DecorProductCatalog.product_id == product_id
        ).first()
        
        if not catalog:
            catalog = DecorProductCatalog(product_id=product_id)
            db.add(catalog)
        
        catalog.cleaning_status = 'repair'
        catalog.product_state = 'damaged'
        catalog.cleaning_last_updated = datetime.now()
        
        # Оновити inventory status
        inventory = db.query(DecorInventoryItem).filter(
            DecorInventoryItem.product_id == product_id
        ).first()
        
        if inventory:
            inventory.status = 'repair'
            inventory.audit_status = 'minor'
        
        # Історія
        history = DecorProductHistory(
            id=f"H-{uuid.uuid4().hex[:8].upper()}",
            product_id=product_id,
            event_type='repair',
            actor=data.get('actor', 'Реквізитор'),
            notes=f"Відправлено на реставрацію: {data.get('notes', '')}"
        )
        db.add(history)
        
        db.commit()
        
        return {
            'success': True,
            'product_id': product_id,
            'message': 'Товар відправлено на реставрацію'
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Помилка: {str(e)}")


@router.post("/items/{item_id}/create-damage-case")
async def create_damage_case(
    item_id: str,
    data: dict,
    db: Session = Depends(get_rh_db)
):
    """
    Створити кейс пошкодження
    """
    try:
        product_id = int(item_id.replace('A-', ''))
        
        # Отримати інформацію про товар
        product = db.query(OpenCartProduct).filter(
            OpenCartProduct.product_id == product_id
        ).first()
        
        if not product:
            raise HTTPException(status_code=404, detail="Товар не знайдено")
        
        description = db.query(OpenCartProductDescription).filter(
            OpenCartProductDescription.product_id == product_id,
            OpenCartProductDescription.language_id == 4
        ).first()
        
        # Знайти останнє замовлення з цим товаром
        inventory = db.query(DecorInventoryItem).filter(
            DecorInventoryItem.product_id == product_id
        ).first()
        
        order_id = inventory.last_order_id if inventory and inventory.last_order_id else None
        
        # Створити кейс пошкодження
        damage_id = f"DMG-{uuid.uuid4().hex[:8].upper()}"
        damage = DecorDamage(
            id=damage_id,
            order_id=order_id,
            order_number=f"ORDER-{order_id}" if order_id else "AUDIT",
            customer_id=0,
            customer_name="Аудит",
            case_status='draft',
            finance_status='none',
            fulfillment_status='none',
            claimed_total=float(data.get('damage_cost', 0)),
            notes=data.get('notes', 'Створено через переоблік'),
            created_by=data.get('actor', 'Реквізитор')
        )
        db.add(damage)
        
        # Створити item в кейсі
        damage_item = DecorDamageItem(
            damage_id=damage_id,
            product_id=product_id,
            barcode=product.model,
            name=description.name if description else product.model,
            image=product.image,
            damage_type=data.get('damage_type', 'physical'),
            qty=1,
            base_value=float(product.ean) if product.ean else 0,
            estimate_value=float(data.get('damage_cost', 0)),
            resolution='pending',
            comment=data.get('notes', '')
        )
        db.add(damage_item)
        
        # Оновити inventory status
        if inventory:
            inventory.audit_status = 'critical'
            inventory.status = 'repair'
        
        # Історія
        history = DecorProductHistory(
            id=f"H-{uuid.uuid4().hex[:8].upper()}",
            product_id=product_id,
            event_type='damage_opened',
            actor=data.get('actor', 'Реквізитор'),
            order_id=order_id,
            notes=f"Відкрито кейс пошкодження: {damage_id}"
        )
        db.add(history)
        
        db.commit()
        
        return {
            'success': True,
            'product_id': product_id,
            'damage_id': damage_id,
            'message': 'Кейс пошкодження створено'
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Помилка: {str(e)}")


@router.put("/items/{item_id}/edit-all")
async def edit_all_item_data(
    item_id: str,
    data: dict,
    db: Session = Depends(get_rh_db)
):
    """
    Редагувати всі дані товару одним запитом
    ✅ MIGRATED: Updates products table (single source of truth)
    """
    try:
        product_id = int(item_id.replace('A-', ''))
        changes = []
        update_fields = {}
        
        # Оновити назву
        if 'name' in data:
            result = db.execute(text("SELECT name FROM products WHERE product_id = :pid"), {"pid": product_id})
            row = result.fetchone()
            if row:
                old_name = row[0]
                update_fields['name'] = data['name']
                changes.append(f"Назва: '{old_name}' → '{data['name']}'")
        
        # Оновити кількість
        if 'qty' in data:
            result = db.execute(text("SELECT quantity FROM products WHERE product_id = :pid"), {"pid": product_id})
            row = result.fetchone()
            if row:
                old_qty = row[0]
                update_fields['quantity'] = data['qty']
                changes.append(f"Кількість: {old_qty} → {data['qty']}")
        
        # ✅ NEW: Оновити колір
        if 'color' in data:
            result = db.execute(text("SELECT color FROM products WHERE product_id = :pid"), {"pid": product_id})
            row = result.fetchone()
            if row:
                old_color = row[0] or '-'
                update_fields['color'] = data['color'] if data['color'] else None
                changes.append(f"Колір: '{old_color}' → '{data['color']}'")
        
        # ✅ NEW: Оновити матеріал
        if 'material' in data:
            result = db.execute(text("SELECT material FROM products WHERE product_id = :pid"), {"pid": product_id})
            row = result.fetchone()
            if row:
                old_material = row[0] or '-'
                update_fields['material'] = data['material'] if data['material'] else None
                changes.append(f"Матеріал: '{old_material}' → '{data['material']}'")
        
        # ✅ NEW: Оновити розмір
        if 'size' in data:
            result = db.execute(text("SELECT size FROM products WHERE product_id = :pid"), {"pid": product_id})
            row = result.fetchone()
            if row:
                old_size = row[0] or '-'
                update_fields['size'] = data['size'] if data['size'] else None
                changes.append(f"Розмір: '{old_size}' → '{data['size']}'")
        
        # Оновити локацію
        if 'zone' in data:
            update_fields['zone'] = data['zone']
            changes.append(f"Зона: {data['zone']}")
        
        if 'location' in data:
            location = data['location']
            update_fields['aisle'] = location[:2] if len(location) >= 2 else location
            update_fields['shelf'] = location[2:] if len(location) > 2 else None
            changes.append(f"Локація: {location}")
        
        # Execute update if there are fields to update
        if update_fields:
            set_clause = ', '.join([f"{k} = :{k}" for k in update_fields.keys()])
            update_fields['pid'] = product_id
            
            db.execute(text(f"""
                UPDATE products 
                SET {set_clause}
                WHERE product_id = :pid
            """), update_fields)
        
        db.commit()
        
        return {
            'success': True,
            'product_id': product_id,
            'changes': changes,
            'message': 'Дані оновлено в products (джерело правди)'
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Помилка: {str(e)}")



@router.get("/categories")
async def get_audit_categories(db: Session = Depends(get_rh_db)):
    """
    ✅ MIGRATED: Отримати категорії з RentalHub DB (categories table)
    """
    try:
        # ✅ NEW: Get categories from RentalHub snapshot table
        main_categories_query = db.execute(text("""
            SELECT DISTINCT name, category_id
            FROM categories
            WHERE parent_id = 0
            ORDER BY name
        """))
        
        main_categories_list = [(row[0], row[1]) for row in main_categories_query if row[0]]
        main_categories = [cat[0] for cat in main_categories_list]
        
        # Create mapping for quick lookup
        cat_id_to_name = {cat_id: cat_name for cat_name, cat_id in main_categories_list}
        
        # Get all subcategories in one query
        if main_categories_list:
            all_subcategories_query = db.execute(text("""
                SELECT parent_id, name
                FROM categories
                WHERE parent_id IN :parent_ids
                ORDER BY parent_id, name
            """), {"parent_ids": tuple([cat_id for _, cat_id in main_categories_list])})
        else:
            all_subcategories_query = []
        
        # Group subcategories by parent
        subcategories_dict = {}
        for parent_id, subcat_name in all_subcategories_query:
            parent_name = cat_id_to_name.get(parent_id)
            if parent_name:
                if parent_name not in subcategories_dict:
                    subcategories_dict[parent_name] = []
                subcategories_dict[parent_name].append(subcat_name)
        
        return {
            'categories': main_categories,
            'subcategories': subcategories_dict
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Помилка отримання категорій: {str(e)}"
        )



@router.post("/mark-category-audited")
async def mark_category_audited(
    data: dict,
    rh_db: Session = Depends(get_rh_db)
):
    """
    Зафіксувати переоблік для всієї категорії
    Оновлює last_audit_date для всіх товарів в категорії
    """
    try:
        category = data.get('category')
        subcategory = data.get('subcategory')
        audited_by = data.get('audited_by', 'Менеджер')
        
        if not category or category == 'all':
            raise HTTPException(
                status_code=400,
                detail="Потрібно вказати конкретну категорію"
            )
        
        # Побудувати умову для запиту
        where_conditions = ["category_name = :category"]
        params = {'category': category, 'audit_date': datetime.now().date()}
        
        if subcategory:
            where_conditions.append("subcategory_name = :subcategory")
            params['subcategory'] = subcategory
        
        where_clause = " AND ".join(where_conditions)
        
        # Оновити всі товари в категорії
        update_query = text(f"""
            UPDATE products 
            SET last_audit_date = :audit_date
            WHERE {where_clause}
        """)
        
        result = rh_db.execute(update_query, params)
        rh_db.commit()
        
        updated_count = result.rowcount
        
        category_name = f"{category}{' · ' + subcategory if subcategory else ''}"
        
        return {
            'success': True,
            'message': f'Категорія "{category_name}" зафіксована як переоблікована',
            'updated_count': updated_count,
            'audit_date': str(datetime.now().date()),
            'audited_by': audited_by
        }
        
    except HTTPException:
        raise
    except Exception as e:
        rh_db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Помилка фіксації переобліку: {str(e)}"
        )



@router.post("/items/{item_id}/toggle-status")
async def toggle_product_status(
    item_id: str,
    data: dict,
    db: Session = Depends(get_rh_db)
):
    """Включити/відключити товар (status 1/0)"""
    try:
        product_id = int(item_id.replace('A-', ''))
        new_status = data.get('status', 0)  # 0 = disabled, 1 = enabled
        db.execute(text("UPDATE products SET status = :st WHERE product_id = :pid"), {"st": new_status, "pid": product_id})
        db.commit()
        label = "включено" if new_status == 1 else "відключено"
        return {"success": True, "message": f"Товар {label}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Помилка: {str(e)}")


@router.delete("/items/{item_id}")
async def delete_product(
    item_id: str,
    db: Session = Depends(get_rh_db)
):
    """Повне видалення товару з БД"""
    try:
        product_id = int(item_id.replace('A-', ''))
        # Видалити пов'язані записи
        db.execute(text("DELETE FROM audit_records WHERE product_id = :pid"), {"pid": product_id})
        db.execute(text("DELETE FROM product_damage_history WHERE product_id = :pid"), {"pid": product_id})
        # Видалити сам товар
        result = db.execute(text("DELETE FROM products WHERE product_id = :pid"), {"pid": product_id})
        db.commit()
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="Товар не знайдено")
        return {"success": True, "message": "Товар видалено"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Помилка: {str(e)}")



@router.get("/export")
async def export_audit_to_excel(
    category: Optional[str] = 'all',
    subcategory: Optional[str] = 'all',
    q: Optional[str] = None,
    rh_db: Session = Depends(get_rh_db)
):
    """
    Експорт даних переобліку в Excel
    Враховує фільтри: категорія, підкатегорія, пошук
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        from fastapi.responses import StreamingResponse
        
        # Побудувати запит з фільтрами
        where_conditions = ["p.status = 1"]
        params = {}
        
        if category and category != 'all':
            where_conditions.append("p.category_name LIKE :category")
            params['category'] = f"%{category}%"
        
        if subcategory and subcategory != 'all':
            where_conditions.append("p.subcategory_name LIKE :subcategory")
            params['subcategory'] = f"%{subcategory}%"
        
        if q:
            where_conditions.append("(p.name LIKE :q OR p.sku LIKE :q OR p.description LIKE :q)")
            params['q'] = f"%{q}%"
        
        where_clause = " AND ".join(where_conditions)
        
        # Запит даних
        query = text(f"""
            SELECT 
                p.product_id,
                p.sku,
                p.name,
                p.category_name,
                p.subcategory_name,
                p.price,
                p.rental_price,
                p.quantity,
                p.color,
                p.material,
                p.size,
                p.zone,
                p.aisle,
                p.shelf,
                p.last_audit_date,
                p.cleaning_status,
                p.product_state,
                p.description,
                p.care_instructions
            FROM products p
            WHERE {where_clause}
            ORDER BY p.category_name, p.subcategory_name, p.name
        """)
        
        results = rh_db.execute(query, params).fetchall()
        
        # Створити Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Переоблік"
        
        # Заголовки
        headers = [
            'ID', 'SKU', 'Назва', 'Категорія', 'Підкатегорія',
            'Ціна купівлі', 'Ціна оренди/день',
            'Кількість', 'Колір', 'Матеріал', 'Розміри',
            'Зона', 'Ряд', 'Полиця', 'Дата переобліку', 
            'Чистота', 'Стан', 'Опис', 'Інструкція по догляду'
        ]
        
        # Стилі заголовків
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Дані
        for row_idx, row_data in enumerate(results, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Автоширина колонок
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Зберегти в пам'ять
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Ім'я файлу
        from datetime import datetime
        from urllib.parse import quote
        filename = f"Audit_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
        filename_encoded = quote(filename)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}"
            }
        )
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Помилка експорту: {str(e)}"
        )


@router.post("/import")
async def import_audit_from_excel(
    file: UploadFile,
    rh_db: Session = Depends(get_rh_db)
):
    """
    Імпорт даних переобліку з Excel
    Оновлює існуючі товари або створює нові
    
    Приймає Excel файл через multipart/form-data
    """
    try:
        from openpyxl import load_workbook
        from io import BytesIO
        
        # Прочитати файл
        contents = await file.read()
        
        # Завантажити Excel
        wb = load_workbook(BytesIO(contents))
        ws = wb.active
        
        updated_count = 0
        created_count = 0
        errors = []
        
        # Пропустити заголовок
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                # Нова структура: ID, SKU, Name, Cat, Subcat, Price, RentalPrice, Qty, Color, Material, Size, Zone, Aisle, Shelf, LastAudit, Cleaning, State, Desc, CareInstr
                if len(row) < 18:
                    errors.append(f"Рядок {row_idx}: недостатньо колонок")
                    continue
                
                product_id, sku, name, category, subcategory, price, rental_price, quantity, color, material, size, zone, aisle, shelf, last_audit_date, cleaning_status, product_state, description, care_instructions = row[:19]
                
                if not sku:
                    continue  # Пропустити порожні рядки
                
                # Перевірити чи існує товар
                check_query = text("SELECT product_id FROM products WHERE sku = :sku")
                existing = rh_db.execute(check_query, {'sku': sku}).fetchone()
                
                if existing:
                    # Оновити існуючий
                    update_query = text("""
                        UPDATE products SET
                            name = :name,
                            category_name = :category,
                            subcategory_name = :subcategory,
                            price = :price,
                            rental_price = :rental_price,
                            quantity = :quantity,
                            color = :color,
                            material = :material,
                            size = :size,
                            zone = :zone,
                            aisle = :aisle,
                            shelf = :shelf,
                            cleaning_status = :cleaning_status,
                            product_state = :product_state,
                            description = :description,
                            care_instructions = :care_instructions
                        WHERE sku = :sku
                    """)
                    
                    rh_db.execute(update_query, {
                        'sku': sku,
                        'name': name,
                        'category': category,
                        'subcategory': subcategory,
                        'price': float(price) if price else 0,
                        'rental_price': float(rental_price) if rental_price else 0,
                        'quantity': quantity or 0,
                        'color': color,
                        'material': material,
                        'size': size,
                        'zone': zone,
                        'aisle': aisle,
                        'shelf': shelf,
                        'cleaning_status': cleaning_status or 'clean',
                        'product_state': product_state or 'good',
                        'description': description,
                        'care_instructions': care_instructions
                    })
                    updated_count += 1
                else:
                    # Створити новий
                    insert_query = text("""
                        INSERT INTO products (
                            product_id, sku, name, category_name, subcategory_name,
                            price, rental_price,
                            quantity, color, material, size,
                            zone, aisle, shelf, cleaning_status, product_state,
                            description, care_instructions, status, last_audit_date
                        ) VALUES (
                            :product_id, :sku, :name, :category, :subcategory,
                            :price, :rental_price,
                            :quantity, :color, :material, :size,
                            :zone, :aisle, :shelf, :cleaning_status, :product_state,
                            :description, :care_instructions, 1, CURDATE()
                        )
                    """)
                    
                    # Генерувати новий product_id
                    max_id_result = rh_db.execute(text("SELECT COALESCE(MAX(product_id), 10000) + 1 FROM products")).fetchone()
                    new_product_id = max_id_result[0]
                    
                    rh_db.execute(insert_query, {
                        'product_id': new_product_id,
                        'sku': sku,
                        'name': name,
                        'category': category,
                        'subcategory': subcategory,
                        'price': float(price) if price else 0,
                        'rental_price': float(rental_price) if rental_price else 0,
                        'quantity': quantity or 0,
                        'color': color,
                        'material': material,
                        'size': size,
                        'zone': zone,
                        'aisle': aisle,
                        'shelf': shelf,
                        'cleaning_status': cleaning_status or 'clean',
                        'product_state': product_state or 'good',
                        'description': description,
                        'care_instructions': care_instructions
                    })
                    created_count += 1
                    
            except Exception as e:
                errors.append(f"Рядок {row_idx}: {str(e)}")
        
        rh_db.commit()
        
        return {
            'success': True,
            'message': f'Імпорт завершено',
            'updated': updated_count,
            'created': created_count,
            'errors': errors[:10]  # Перші 10 помилок
        }
        
    except Exception as e:
        rh_db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Помилка імпорту: {str(e)}"
        )

        
        category_name = f"{category}{' · ' + subcategory if subcategory else ''}"
        
        return {
            'success': True,
            'message': f'Категорія "{category_name}" зафіксована як переоблікована',
            'updated_count': updated_count,
            'audit_date': str(datetime.now().date()),
            'audited_by': audited_by
        }
        
    except HTTPException:
        raise
    except Exception as e:
        rh_db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Помилка фіксації переобліку: {str(e)}"
        )

